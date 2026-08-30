"""
tests/test_reconciliation.py
-----------------------------
Unit tests for every pure rule function in app.services.reconciliation.

Design principles
-----------------
* Each test builds the **minimum DataFrame** needed to exercise one rule.
* Tests are independent — no shared state, no file I/O, no DB.
* Pure functions are imported directly so they can be tested without the
  ReconciliationEngine orchestrator class.
* Edge cases (NaN, zero amounts, boundary dates) are tested alongside the
  happy/unhappy paths.

Run with:
    cd backend
    pytest tests/test_reconciliation.py -v
"""

from __future__ import annotations

import pandas as pd
import pytest
from datetime import datetime, timedelta, timezone

from app.services.reconciliation import (
    ExcType,
    build_result,
    classify,
    load_payments,
    load_settlements,
    load_bank_statement,
    merge_all,
    rule_amount_mismatch,
    rule_delayed_settlement,
    rule_duplicate_payments,
    rule_unmatched_no_bank_credit,
    rule_unmatched_no_settlement,
)


# ─────────────────────────────────────────────────────────────────────────────
# Factory helpers — build minimal DataFrames for one rule at a time
# ─────────────────────────────────────────────────────────────────────────────

_NOW = datetime(2024, 1, 10, 12, 0, 0)


def _payment_row(**overrides) -> dict:
    """Return a single payment dict with sensible defaults."""
    base = {
        "payment_id":   "pay_001",
        "order_id":     "ord_001",
        "amount":       1000.0,
        "currency":     "INR",
        "payment_date": _NOW,
        "method":       "upi",
        "status":       "captured",
        "utr_number":   "UTR001",
    }
    base.update(overrides)
    return base


def _settlement_row(**overrides) -> dict:
    """Return a single settlement dict with sensible defaults."""
    base = {
        "settlement_id":   "setl_001",
        "payment_id":      "pay_001",
        "settled_amount":  1000.0,
        "settlement_date": _NOW + timedelta(days=1),
        "utr_number":      "UTR001",
    }
    base.update(overrides)
    return base


def _bank_row(**overrides) -> dict:
    """Return a single bank-statement dict with sensible defaults."""
    base = {
        "utr_number":       "UTR001",
        "bank_credited_amount": 1000.0,
        "bank_credit_date": _NOW + timedelta(days=1),
        "bank_narration":   "NEFT/UTR001/settlement",
    }
    base.update(overrides)
    return base


def _merge_one(payment: dict, settlement: dict | None, bank: dict | None) -> pd.DataFrame:
    """
    Build a single-row merged DataFrame from component dicts.
    Passing None for settlement/bank simulates a missing join partner.
    """
    pay_df  = pd.DataFrame([payment])
    setl_df = pd.DataFrame([settlement]) if settlement else pd.DataFrame(
        columns=["settlement_id", "payment_id", "settled_amount", "settlement_date", "utr_number"]
    )
    bank_df = pd.DataFrame([bank]) if bank else pd.DataFrame(
        columns=["utr_number", "bank_credited_amount", "bank_credit_date", "bank_narration"]
    )
    return merge_all(pay_df, setl_df, bank_df)


# ─────────────────────────────────────────────────────────────────────────────
# rule_unmatched_no_settlement
# ─────────────────────────────────────────────────────────────────────────────

class TestRuleUnmatchedNoSettlement:
    """
    Business rule: a payment with no settlement record is critical — funds
    may be stuck with the acquirer.
    """

    def test_flags_when_settlement_missing(self):
        df = _merge_one(_payment_row(), settlement=None, bank=None)
        mask = rule_unmatched_no_settlement(df)
        assert mask.all(), "Should flag payment with no settlement"

    def test_does_not_flag_when_settlement_present(self):
        df = _merge_one(_payment_row(), _settlement_row(), _bank_row())
        mask = rule_unmatched_no_settlement(df)
        assert not mask.any(), "Should NOT flag when settlement row exists"

    def test_flags_only_the_missing_row_in_mixed_df(self):
        p1 = _payment_row(payment_id="pay_001", utr_number="UTR001")
        p2 = _payment_row(payment_id="pay_002", utr_number="UTR002")
        s1 = _settlement_row(payment_id="pay_001")

        pay_df  = pd.DataFrame([p1, p2])
        setl_df = pd.DataFrame([s1])
        bank_df = pd.DataFrame(
            columns=["utr_number", "bank_credited_amount", "bank_credit_date", "bank_narration"]
        )
        df   = merge_all(pay_df, setl_df, bank_df)
        mask = rule_unmatched_no_settlement(df)

        assert mask.sum() == 1, "Only pay_002 should be flagged"
        assert df.loc[mask, "payment_id"].iloc[0] == "pay_002"


# ─────────────────────────────────────────────────────────────────────────────
# rule_unmatched_no_bank_credit
# ─────────────────────────────────────────────────────────────────────────────

class TestRuleUnmatchedNoBankCredit:
    """
    Business rule: settlement exists in the internal system but no matching
    bank credit was found — the transfer may have failed.
    """

    def test_flags_when_bank_credit_missing(self):
        df = _merge_one(_payment_row(), _settlement_row(), bank=None)
        mask = rule_unmatched_no_bank_credit(df)
        assert mask.all(), "Should flag settlement with no bank credit"

    def test_does_not_flag_when_bank_credit_present(self):
        df = _merge_one(_payment_row(), _settlement_row(), _bank_row())
        mask = rule_unmatched_no_bank_credit(df)
        assert not mask.any()

    def test_does_not_flag_when_no_settlement_either(self):
        """
        If there is no settlement at all, this rule should not trigger —
        that case belongs to rule_unmatched_no_settlement.
        """
        df = _merge_one(_payment_row(), settlement=None, bank=None)
        mask = rule_unmatched_no_bank_credit(df)
        assert not mask.any(), "No settlement → no bank credit rule should stay silent"


# ─────────────────────────────────────────────────────────────────────────────
# rule_amount_mismatch
# ─────────────────────────────────────────────────────────────────────────────

class TestRuleAmountMismatch:
    """
    Business rule: |payment - settled| > max(₹1, 0.5 % × payment) is flagged.
    Small processing-fee differences below the tolerance band are ignored.
    """

    def test_flags_when_difference_exceeds_absolute_tolerance(self):
        df = _merge_one(
            _payment_row(amount=1000.0),
            _settlement_row(settled_amount=998.0),   # Δ = 2 > max(1, 5) = 5? No...
            _bank_row(),
        )
        # tolerance = max(1, 0.005*1000) = max(1, 5) = 5
        # Δ = 2 < 5 → should NOT flag
        mask = rule_amount_mismatch(df, tolerance_abs=1.0, tolerance_pct=0.005)
        assert not mask.any(), "Δ=2 is inside 0.5% tolerance band of ₹1000, should not flag"

    def test_flags_when_difference_exceeds_pct_tolerance(self):
        df = _merge_one(
            _payment_row(amount=1000.0),
            _settlement_row(settled_amount=992.0),   # Δ = 8 > max(1, 5) = 5 → flag
            _bank_row(),
        )
        mask = rule_amount_mismatch(df, tolerance_abs=1.0, tolerance_pct=0.005)
        assert mask.all(), "Δ=8 exceeds 0.5% band (₹5), should flag"

    def test_flags_small_amount_exceeding_abs_tolerance(self):
        """For tiny payments the absolute floor (₹1) is the operative threshold."""
        df = _merge_one(
            _payment_row(amount=10.0),
            _settlement_row(settled_amount=8.5),     # Δ = 1.5 > max(1, 0.05) = 1 → flag
            _bank_row(),
        )
        mask = rule_amount_mismatch(df, tolerance_abs=1.0, tolerance_pct=0.005)
        assert mask.all(), "Δ=1.5 exceeds abs floor of ₹1 for a ₹10 payment"

    def test_does_not_flag_exact_match(self):
        df = _merge_one(_payment_row(amount=500.0), _settlement_row(settled_amount=500.0), _bank_row())
        mask = rule_amount_mismatch(df)
        assert not mask.any()

    def test_does_not_flag_when_no_settlement(self):
        df = _merge_one(_payment_row(), settlement=None, bank=None)
        mask = rule_amount_mismatch(df)
        assert not mask.any(), "No settlement → mismatch rule should stay silent"

    def test_custom_tolerance_respected(self):
        df = _merge_one(
            _payment_row(amount=100.0),
            _settlement_row(settled_amount=97.0),   # Δ = 3
            _bank_row(),
        )
        # With a wide tolerance of ₹5 absolute, Δ=3 should not flag
        assert not rule_amount_mismatch(df, tolerance_abs=5.0, tolerance_pct=0.0).any()
        # With a tight tolerance of ₹2 absolute, Δ=3 should flag
        assert rule_amount_mismatch(df, tolerance_abs=2.0, tolerance_pct=0.0).all()


# ─────────────────────────────────────────────────────────────────────────────
# rule_delayed_settlement
# ─────────────────────────────────────────────────────────────────────────────

class TestRuleDelayedSettlement:
    """
    Business rule: settlement received after T+{sla_days} calendar days is
    flagged as DELAYED_SETTLEMENT.  Default SLA is T+2.
    """

    def test_flags_t_plus_3_with_default_sla(self):
        pay  = _payment_row(payment_date=_NOW)
        setl = _settlement_row(settlement_date=_NOW + timedelta(days=3))
        df   = _merge_one(pay, setl, _bank_row())
        mask = rule_delayed_settlement(df, sla_days=2)
        assert mask.all(), "T+3 exceeds T+2 SLA, should flag"

    def test_does_not_flag_t_plus_2(self):
        pay  = _payment_row(payment_date=_NOW)
        setl = _settlement_row(settlement_date=_NOW + timedelta(days=2))
        df   = _merge_one(pay, setl, _bank_row())
        mask = rule_delayed_settlement(df, sla_days=2)
        assert not mask.any(), "T+2 is exactly the SLA, should NOT flag"

    def test_does_not_flag_t_plus_1(self):
        pay  = _payment_row(payment_date=_NOW)
        setl = _settlement_row(settlement_date=_NOW + timedelta(days=1))
        df   = _merge_one(pay, setl, _bank_row())
        mask = rule_delayed_settlement(df, sla_days=2)
        assert not mask.any()

    def test_custom_sla_respected(self):
        pay  = _payment_row(payment_date=_NOW)
        setl = _settlement_row(settlement_date=_NOW + timedelta(days=5))
        df   = _merge_one(pay, setl, _bank_row())
        assert rule_delayed_settlement(df, sla_days=4).all()   # 5 > 4 → flag
        assert not rule_delayed_settlement(df, sla_days=5).any()  # 5 == 5 → ok

    def test_does_not_flag_when_no_settlement(self):
        df = _merge_one(_payment_row(), settlement=None, bank=None)
        mask = rule_delayed_settlement(df)
        assert not mask.any()


# ─────────────────────────────────────────────────────────────────────────────
# rule_duplicate_payments
# ─────────────────────────────────────────────────────────────────────────────

class TestRuleDuplicatePayments:
    """
    Business rule: two captured payments for the same (order_id, amount)
    within 60 seconds are flagged as DUPLICATE.
    """

    def _payments_df(self, rows: list[dict]) -> pd.DataFrame:
        """Build a payments DataFrame from a list of dicts and add a dummy merged frame."""
        df = pd.DataFrame(rows)
        df["payment_date"]   = pd.to_datetime(df["payment_date"])
        df["settlement_id"]  = None           # simulate no-settlement rows
        df["settled_amount"] = None
        df["utr_number"]     = None
        df["bank_credited_amount"] = None
        return df

    def test_flags_two_payments_within_window(self):
        t0 = _NOW
        df = self._payments_df([
            _payment_row(payment_id="pay_001", payment_date=t0),
            _payment_row(payment_id="pay_002", payment_date=t0 + timedelta(seconds=30)),
        ])
        mask = rule_duplicate_payments(df, window_secs=300)
        assert mask.sum() == 2, "Both pay_001 and pay_002 should be flagged"

    def test_does_not_flag_payments_outside_window(self):
        t0 = _NOW
        df = self._payments_df([
            _payment_row(payment_id="pay_001", payment_date=t0),
            _payment_row(payment_id="pay_002", payment_date=t0 + timedelta(seconds=360)),
        ])
        mask = rule_duplicate_payments(df, window_secs=300)
        assert not mask.any(), "360 s apart is outside 300 s window, should not flag"

    def test_does_not_flag_different_orders(self):
        t0 = _NOW
        df = self._payments_df([
            _payment_row(payment_id="pay_001", order_id="ord_001", payment_date=t0),
            _payment_row(payment_id="pay_002", order_id="ord_002", payment_date=t0 + timedelta(seconds=5)),
        ])
        mask = rule_duplicate_payments(df, window_secs=60)
        assert not mask.any(), "Different orders should not be flagged"

    def test_does_not_flag_different_amounts(self):
        """Same order, same time, but different amounts — NOT a duplicate."""
        t0 = _NOW
        df = self._payments_df([
            _payment_row(payment_id="pay_001", order_id="ord_001", amount=500.0, payment_date=t0),
            _payment_row(payment_id="pay_002", order_id="ord_001", amount=300.0, payment_date=t0 + timedelta(seconds=5)),
        ])
        mask = rule_duplicate_payments(df, window_secs=60)
        assert not mask.any(), "Different amounts in same order are not duplicates"

    def test_does_not_flag_failed_payments(self):
        """Only 'captured' payments are in scope."""
        t0 = _NOW
        df = self._payments_df([
            _payment_row(payment_id="pay_001", status="failed",   payment_date=t0),
            _payment_row(payment_id="pay_002", status="captured", payment_date=t0 + timedelta(seconds=5)),
        ])
        mask = rule_duplicate_payments(df, window_secs=60)
        assert not mask.any(), "Failed payment should not be considered for duplicate check"

    def test_flags_only_close_pair_in_larger_group(self):
        """In a group of 3, only the pair within the window should be flagged."""
        t0 = _NOW
        df = self._payments_df([
            _payment_row(payment_id="pay_001", payment_date=t0),
            _payment_row(payment_id="pay_002", payment_date=t0 + timedelta(seconds=30)),
            _payment_row(payment_id="pay_003", payment_date=t0 + timedelta(hours=2)),
        ])
        mask = rule_duplicate_payments(df, window_secs=300)
        flagged_ids = set(df.loc[mask, "payment_id"].tolist())
        assert flagged_ids == {"pay_001", "pay_002"}, "Only the close pair should be flagged"
        assert "pay_003" not in flagged_ids

    def test_flags_payments_3_minutes_apart(self):
        """
        Regression test for the original 60 s window bug.
        The sample data generator injects duplicates 1-3 minutes apart
        (timedelta(minutes=randint(1, 3))).  The engine must catch the
        worst-case 3-minute (180 s) gap with the 300 s default window.
        """
        t0 = _NOW
        df = self._payments_df([
            _payment_row(payment_id="pay_original", payment_date=t0),
            _payment_row(payment_id="pay_retry",    payment_date=t0 + timedelta(minutes=3)),
        ])
        # With default 300 s window: 180 s < 300 s → must flag
        mask = rule_duplicate_payments(df)           # uses default window_secs=300
        assert mask.sum() == 2, (
            "3-minute gap (180 s) is within the 300 s window — both rows must be flagged. "
            "This catches the case that was missed when window_secs=60."
        )

    def test_does_not_flag_at_exactly_boundary(self):
        """Payments exactly at the boundary (300 s) are still flagged (≤, not <)."""
        t0 = _NOW
        df = self._payments_df([
            _payment_row(payment_id="pay_001", payment_date=t0),
            _payment_row(payment_id="pay_002", payment_date=t0 + timedelta(seconds=300)),
        ])
        mask = rule_duplicate_payments(df)
        assert mask.sum() == 2, "Exactly at boundary (300 s) should still be flagged (≤)"

    def test_bank_join_duplicate_rows_do_not_create_false_positives(self):
        """
        Regression test for the bank-join artifact bug.

        When a UTR appears more than once in bank_statement.csv (e.g. the
        generator's deliberate 'bonus' duplicate credit entry), the merged
        DataFrame can contain two rows for the same payment_id.  Without
        `drop_duplicates(subset=['payment_id'])` the groupby would see:

            (order_id, amount) group = [pay_X(t0), pay_X(t0)]   ← same payment, 2 rows

        That looks like a 0-second consecutive pair → falsely flagged DUPLICATE.

        This test simulates the artifact by injecting a repeated payment_id row
        and asserts the rule returns ZERO false positives.  Separately, it
        verifies that a genuine duplicate pair (2 distinct payment_ids within
        the window) is still correctly flagged, so the fix doesn't over-suppress.
        """
        t0 = _NOW

        # --- 1. Genuine duplicate pair (should be flagged) ---
        genuine_original = _payment_row(
            payment_id="pay_orig", order_id="ord_A",
            amount=500.0, payment_date=t0,
        )
        genuine_retry = _payment_row(
            payment_id="pay_retry", order_id="ord_A",
            amount=500.0, payment_date=t0 + timedelta(seconds=90),
        )

        # --- 2. Non-duplicate payment whose bank join created 2 rows ---
        solo_payment = _payment_row(
            payment_id="pay_solo", order_id="ord_B",
            amount=1000.0, payment_date=t0,
        )

        df = self._payments_df([genuine_original, genuine_retry, solo_payment, solo_payment])
        # ^ solo_payment appears TWICE, simulating the bank join artifact

        mask = rule_duplicate_payments(df, window_secs=300)

        flagged_ids = set(df.loc[mask, "payment_id"].tolist())

        # Exactly 2 rows flagged from the genuine pair (pay_orig + pay_retry)
        assert flagged_ids == {"pay_orig", "pay_retry"}, (
            f"Expected only genuine pair flagged. Got: {flagged_ids}. "
            "Bank-join-duplicated payment_id must not count as a self-duplicate."
        )
        assert mask.sum() == 2, (
            f"Expected 2 flagged rows, got {mask.sum()}. "
            "The extra bank row must not inflate the count."
        )



# ─────────────────────────────────────────────────────────────────────────────
# classify — priority ordering integration test
# ─────────────────────────────────────────────────────────────────────────────

class TestClassifyPriority:
    """
    When multiple rules match the same row, higher-priority rules must win.
    Updated priority (highest → lowest):
        DUPLICATE > UNMATCHED_NO_SETTLEMENT > UNMATCHED_NO_BANK_CREDIT >
        AMOUNT_MISMATCH > DELAYED_SETTLEMENT
    """

    def test_all_matched(self):
        df = _merge_one(_payment_row(), _settlement_row(), _bank_row())
        out = classify(df)
        assert out["recon_status"].iloc[0] == ExcType.MATCHED

    def test_unmatched_no_settlement_wins_over_delayed(self):
        """
        A payment with no settlement and no duplicate pattern should be
        UNMATCHED_NO_SETTLEMENT, not DELAYED_SETTLEMENT.
        DUPLICATE is the only rule that outranks UNMATCHED_NO_SETTLE.
        """
        df = _merge_one(_payment_row(), settlement=None, bank=None)
        out = classify(df, sla_days=0)
        assert out["recon_status"].iloc[0] == ExcType.UNMATCHED_NO_SETTLE

    def test_duplicate_wins_over_unmatched_no_settlement(self):
        """
        DUPLICATE has highest priority and overwrites UNMATCHED_NO_SETTLEMENT.

        Real-world scenario: duplicate payments are typically rejected by the
        processor so they never get a settlement.  Without this rule the engine
        would classify them as UNMATCHED_NO_SETTLEMENT, hiding the true cause.
        """
        t0 = _NOW
        # Two payments: same order_id, same amount, 30 s apart but NO settlement.
        p1 = _payment_row(payment_id="pay_original", payment_date=t0)
        p2 = _payment_row(payment_id="pay_retry",    payment_date=t0 + timedelta(seconds=30))

        pay_df  = pd.DataFrame([p1, p2])
        pay_df["payment_date"] = pd.to_datetime(pay_df["payment_date"])
        setl_df = pd.DataFrame(
            columns=["settlement_id", "payment_id", "settled_amount", "settlement_date", "utr_number"]
        )
        bank_df = pd.DataFrame(
            columns=["utr_number", "bank_credited_amount", "bank_credit_date", "bank_narration"]
        )
        merged = merge_all(pay_df, setl_df, bank_df)
        out = classify(merged, dup_window_secs=300)

        # Both rows must be DUPLICATE, not UNMATCHED_NO_SETTLEMENT
        statuses = set(out["recon_status"].tolist())
        assert statuses == {ExcType.DUPLICATE}, (
            f"Expected all rows = DUPLICATE, got {statuses}. "
            "DUPLICATE must outrank UNMATCHED_NO_SETTLEMENT."
        )

    def test_unmatched_no_bank_wins_over_delayed(self):
        """Settlement exists but bank is missing → NO_BANK_CREDIT, not DELAYED."""
        pay  = _payment_row(payment_date=_NOW)
        setl = _settlement_row(settlement_date=_NOW + timedelta(days=10))
        df   = _merge_one(pay, setl, bank=None)
        out  = classify(df, sla_days=2)
        assert out["recon_status"].iloc[0] == ExcType.UNMATCHED_NO_BANK

    def test_amount_mismatch_wins_over_delayed(self):
        pay  = _payment_row(amount=1000.0, payment_date=_NOW)
        setl = _settlement_row(settled_amount=900.0, settlement_date=_NOW + timedelta(days=5))
        df   = _merge_one(pay, setl, _bank_row())
        out  = classify(df, sla_days=2)
        assert out["recon_status"].iloc[0] == ExcType.AMOUNT_MISMATCH

    def test_delay_days_column_added(self):
        pay  = _payment_row(payment_date=_NOW)
        setl = _settlement_row(settlement_date=_NOW + timedelta(days=3))
        df   = _merge_one(pay, setl, _bank_row())
        out  = classify(df, sla_days=2)
        assert "delay_days" in out.columns
        assert out["delay_days"].iloc[0] == 3


# ─────────────────────────────────────────────────────────────────────────────
# build_result — output schema
# ─────────────────────────────────────────────────────────────────────────────

class TestBuildResult:
    """Verify the result dict shape and computed values."""

    def _make_classified(self, statuses: list[str]) -> pd.DataFrame:
        """Create a dummy classified DataFrame with given statuses."""
        rows = []
        for i, status in enumerate(statuses):
            rows.append({
                "payment_id":    f"pay_{i:03d}",
                "order_id":      f"ord_{i:03d}",
                "amount":        float(100 * (i + 1)),
                "settled_amount": float(100 * (i + 1)) if status == ExcType.MATCHED else float(100 * (i + 1) - 10),
                "settlement_id":  f"setl_{i}" if status != ExcType.UNMATCHED_NO_SETTLE else None,
                "utr_number":     f"UTR{i:03d}" if status not in (ExcType.UNMATCHED_NO_SETTLE, ExcType.UNMATCHED_NO_BANK) else None,
                "bank_credited_amount": None,
                "payment_date":   _NOW,
                "settlement_date": _NOW + timedelta(days=1),
                "delay_days":     1,
                "recon_status":   status,
                "details":        "test exception",
                "severity":       "medium",
            })
        return pd.DataFrame(rows)

    def test_100_percent_match_rate(self):
        df = self._make_classified([ExcType.MATCHED, ExcType.MATCHED])
        result = build_result(df, sla_days=2)
        assert result["match_rate"] == 100.0
        assert result["matched"] == 2
        assert result["unmatched"] == 0
        assert result["exceptions"] == []

    def test_zero_percent_match_rate(self):
        df = self._make_classified([ExcType.UNMATCHED_NO_SETTLE, ExcType.AMOUNT_MISMATCH])
        result = build_result(df, sla_days=2)
        assert result["match_rate"] == 0.0
        assert result["matched"] == 0
        assert result["unmatched"] == 2
        assert len(result["exceptions"]) == 2

    def test_mixed_match_rate(self):
        statuses = [ExcType.MATCHED] * 3 + [ExcType.AMOUNT_MISMATCH]
        df = self._make_classified(statuses)
        result = build_result(df, sla_days=2)
        assert result["match_rate"] == 75.0
        assert result["matched"] == 3
        assert result["total_records"] == 4

    def test_required_keys_present(self):
        df = self._make_classified([ExcType.MATCHED])
        result = build_result(df, sla_days=2)
        required_keys = {
            "match_rate", "total_records", "matched", "unmatched",
            "exception_counts", "exceptions", "settlement_totals",
            "sla_days_used", "run_at",
        }
        assert required_keys.issubset(result.keys())

    def test_exception_counts_aggregated_by_type(self):
        statuses = [
            ExcType.AMOUNT_MISMATCH,
            ExcType.AMOUNT_MISMATCH,
            ExcType.DELAYED_SETTLEMENT,
        ]
        df = self._make_classified(statuses)
        result = build_result(df, sla_days=2)
        counts = result["exception_counts"]
        assert counts[ExcType.AMOUNT_MISMATCH]   == 2
        assert counts[ExcType.DELAYED_SETTLEMENT] == 1

    def test_sla_days_echoed_in_result(self):
        df = self._make_classified([ExcType.MATCHED])
        assert build_result(df, sla_days=5)["sla_days_used"] == 5

    def test_settlement_totals_keys_present(self):
        df = self._make_classified([ExcType.MATCHED])
        totals = build_result(df, sla_days=2)["settlement_totals"]
        assert {"total_payment_amount", "total_settled_amount", "currency"}.issubset(totals.keys())

    def test_settlement_totals_split_figures_computed(self):
        statuses = [
            ExcType.MATCHED,
            ExcType.AMOUNT_MISMATCH,
            ExcType.UNMATCHED_NO_SETTLE,
            ExcType.DUPLICATE,
        ]
        df = self._make_classified(statuses)
        # Set unrealistic settled amounts to None for rows that should have 0 settled
        # (UNMATCHED_NO_SETTLE and DUPLICATE that were rejected)
        df.loc[df["recon_status"] == ExcType.UNMATCHED_NO_SETTLE, "settled_amount"] = None
        df.loc[df["recon_status"] == ExcType.DUPLICATE, "settled_amount"] = 0.0  # Duplicate was rejected
        
        result = build_result(df, sla_days=2)
        totals = result["settlement_totals"]
        assert "amount_mismatches_total" in totals
        assert "unsettled_value_total" in totals
        assert "duplicate_charges_total" in totals
        assert totals["amount_mismatches_total"] > 0
        assert totals["unsettled_value_total"] > 0
        assert totals["duplicate_charges_total"] > 0
        # With the corrected data:
        # AMOUNT_MISMATCH: |200 - 190| = 10
        # UNMATCHED_NO_SETTLE: |300 - 0| = 300 (settled_amount was set to None, treated as 0)
        # DUPLICATE: |400 - 0| = 400 (settled_amount = 0, duplicate rejected)
        assert totals["amount_mismatches_total"] == 10.0
        assert totals["unsettled_value_total"] == 300.0
        assert totals["duplicate_charges_total"] == 400.0

    def test_settlement_totals_sum_invariant(self):
        """Confirm Amount Mismatches + Unsettled Value + Duplicate Charges == Gross Discrepancy Total exactly."""
        statuses = [
            ExcType.MATCHED,
            ExcType.AMOUNT_MISMATCH,
            ExcType.UNMATCHED_NO_SETTLE,
            ExcType.DUPLICATE,
        ]
        df = self._make_classified(statuses)
        # AMOUNT_MISMATCH: amount 200, settled 190 -> diff 10
        # UNMATCHED_NO_SETTLE: amount 300, settled 0 -> diff 300
        # DUPLICATE: amount 400, settled 0 (unsettled duplicate) -> diff 400
        df.loc[df["recon_status"] == ExcType.UNMATCHED_NO_SETTLE, "settled_amount"] = None
        df.loc[df["recon_status"] == ExcType.DUPLICATE, "settled_amount"] = None

        result = build_result(df, sla_days=2)
        totals = result["settlement_totals"]
        mismatch_sum = totals["amount_mismatches_total"]
        unsettled_sum = totals["unsettled_value_total"]
        duplicate_sum = totals["duplicate_charges_total"]
        gross_discrepancy = totals["total_discrepancy"]

        assert round(mismatch_sum + unsettled_sum + duplicate_sum, 2) == round(gross_discrepancy, 2)

    def test_settlement_totals_sum_invariant_with_realistic_duplicates(self):
        """
        Test sum invariant with realistic DUPLICATE scenario:
        - Some duplicate pairs are BOTH settled in bank (discrepancy = 0)
        - Some duplicate pairs have only ONE settled in bank, not the duplicate (discrepancy = amount)
        
        This validates that duplicate_charges_total uses discrepancy_amount, not internal_amount.
        Regression test for: https://github.com/user/razorrecon/issues/X
        """
        # Build DataFrame with:
        # - 1 MATCHED payment
        # - 1 AMOUNT_MISMATCH (amount=100, settled=90, discrepancy=10)
        # - 3 DUPLICATE payments:
        #   - pay_A1 and pay_A2: both amount=50, both settled=50, discrepancy=0 (both credited to bank)
        #   - pay_B1 and pay_B2: both amount=60, B1 settled=60, B2 settled=0, B2 discrepancy=60 (one rejected)
        # - 1 UNMATCHED_NO_SETTLE: amount=200, settled=0, discrepancy=200
        #
        # Total discrepancy should be: 10 + 0 + 0 + 60 + 200 = 270
        # 3-way split should be:
        #   - Amount Mismatches: 10
        #   - Unsettled Value: 200
        #   - Duplicate Charges: 60 (not 110, which would be 50+60)
        
        df = pd.DataFrame([
            # MATCHED
            {"payment_id": "p_match", "order_id": "o_001", "amount": 1000.0, "settled_amount": 1000.0, "recon_status": ExcType.MATCHED},
            # AMOUNT_MISMATCH
            {"payment_id": "p_amm", "order_id": "o_002", "amount": 100.0, "settled_amount": 90.0, "recon_status": ExcType.AMOUNT_MISMATCH},
            # DUPLICATE pair 1: both settled (discrepancy = 0 for both)
            {"payment_id": "p_a1", "order_id": "o_003", "amount": 50.0, "settled_amount": 50.0, "recon_status": ExcType.DUPLICATE},
            {"payment_id": "p_a2", "order_id": "o_003", "amount": 50.0, "settled_amount": 50.0, "recon_status": ExcType.DUPLICATE},
            # DUPLICATE pair 2: one settled, one not (discrepancy = 60 for the unsettled one)
            {"payment_id": "p_b1", "order_id": "o_004", "amount": 60.0, "settled_amount": 60.0, "recon_status": ExcType.DUPLICATE},
            {"payment_id": "p_b2", "order_id": "o_004", "amount": 60.0, "settled_amount": 0.0, "recon_status": ExcType.DUPLICATE},
            # UNMATCHED_NO_SETTLE
            {"payment_id": "p_uns", "order_id": "o_005", "amount": 200.0, "settled_amount": 0.0, "recon_status": ExcType.UNMATCHED_NO_SETTLE},
        ])
        
        result = build_result(df, sla_days=2)
        totals = result["settlement_totals"]
        
        mismatch_sum = totals["amount_mismatches_total"]
        unsettled_sum = totals["unsettled_value_total"]
        duplicate_sum = totals["duplicate_charges_total"]
        gross_discrepancy = totals["total_discrepancy"]
        
        # Verify each bucket
        assert round(mismatch_sum, 2) == 10.0, f"Amount mismatch should be 10, got {mismatch_sum}"
        assert round(unsettled_sum, 2) == 200.0, f"Unsettled should be 200, got {unsettled_sum}"
        assert round(duplicate_sum, 2) == 60.0, f"Duplicate should be 60 (only unsettled duplicate), got {duplicate_sum}"
        
        # Total should be 10 + 200 + 60 = 270
        assert round(mismatch_sum + unsettled_sum + duplicate_sum, 2) == 270.0
        assert round(gross_discrepancy, 2) == 270.0
        
        # Most importantly: the sum invariant must hold
        assert round(mismatch_sum + unsettled_sum + duplicate_sum, 2) == round(gross_discrepancy, 2), \
            f"Sum invariant FAILED: {mismatch_sum} + {unsettled_sum} + {duplicate_sum} = {mismatch_sum + unsettled_sum + duplicate_sum} != {gross_discrepancy}"


# ─────────────────────────────────────────────────────────────────────────────
# Exception structured fields — payment_id, order_id, exception_date
# ─────────────────────────────────────────────────────────────────────────────

class TestExceptionStructuredFields:
    """
    Every exception dict must expose payment_id, order_id, and exception_date
    as first-class structured fields.

    exception_date per type:
        UNMATCHED_NO_SETTLEMENT  -> payment_date
        DUPLICATE                -> payment_date
        AMOUNT_MISMATCH          -> settlement_date
        UNMATCHED_NO_BANK_CREDIT -> settlement_date
        DELAYED_SETTLEMENT       -> settlement_date
    """
    PAY_DATE  = _NOW
    SETL_DATE = _NOW + timedelta(days=3)

    def _first_exc(self, exc_type, payment=None, settlement=None, bank=None):
        pay  = payment or _payment_row(payment_date=self.PAY_DATE)
        df   = _merge_one(pay, settlement, bank)
        excs = [e for e in build_result(classify(df), sla_days=2)["exceptions"]
                if e["type"] == exc_type]
        assert excs, f"No exception of type {exc_type} generated"
        return excs[0]

    def test_unmatched_no_settlement_uses_payment_date(self):
        exc = self._first_exc(ExcType.UNMATCHED_NO_SETTLE)
        assert exc["payment_id"]     == "pay_001"
        assert exc["order_id"]       == "ord_001"
        assert exc["exception_date"] is not None
        assert "2024-01-10" in exc["exception_date"]

    def test_duplicate_uses_payment_date(self):
        t0 = self.PAY_DATE
        p1 = _payment_row(payment_id="pay_001", payment_date=t0)
        p2 = _payment_row(payment_id="pay_002", payment_date=t0 + timedelta(seconds=30))
        pay_df  = pd.DataFrame([p1, p2])
        pay_df["payment_date"] = pd.to_datetime(pay_df["payment_date"])
        setl_df = pd.DataFrame(columns=["settlement_id","payment_id","settled_amount","settlement_date","utr_number"])
        bank_df = pd.DataFrame(columns=["utr_number","bank_credited_amount","bank_credit_date","bank_narration"])
        merged  = merge_all(pay_df, setl_df, bank_df)
        result  = build_result(classify(merged, dup_window_secs=300), sla_days=2)
        dup_excs = [e for e in result["exceptions"] if e["type"] == ExcType.DUPLICATE]
        assert len(dup_excs) == 2
        for exc in dup_excs:
            assert exc["payment_id"]     is not None
            assert exc["order_id"]       == "ord_001"
            assert exc["exception_date"] is not None
            assert "2024-01-10" in exc["exception_date"]

    def test_amount_mismatch_uses_settlement_date(self):
        pay  = _payment_row(amount=1000.0, payment_date=self.PAY_DATE)
        setl = _settlement_row(settled_amount=980.0, settlement_date=self.SETL_DATE)
        exc  = self._first_exc(ExcType.AMOUNT_MISMATCH, pay, setl, _bank_row())
        assert exc["payment_id"] == "pay_001"
        assert exc["order_id"]   == "ord_001"
        expected = pd.Timestamp(self.SETL_DATE).isoformat()[:10]
        assert expected in exc["exception_date"]

    def test_no_bank_credit_uses_settlement_date(self):
        pay  = _payment_row(payment_date=self.PAY_DATE)
        setl = _settlement_row(settlement_date=self.SETL_DATE)
        exc  = self._first_exc(ExcType.UNMATCHED_NO_BANK, pay, setl, bank=None)
        assert exc["payment_id"] == "pay_001"
        assert exc["order_id"]   == "ord_001"
        expected = pd.Timestamp(self.SETL_DATE).isoformat()[:10]
        assert expected in exc["exception_date"]

    def test_delayed_settlement_uses_settlement_date(self):
        pay  = _payment_row(payment_date=self.PAY_DATE)
        setl = _settlement_row(settlement_date=self.SETL_DATE)
        exc  = self._first_exc(ExcType.DELAYED_SETTLEMENT, pay, setl, _bank_row())
        assert exc["payment_id"] == "pay_001"
        assert exc["order_id"]   == "ord_001"
        expected = pd.Timestamp(self.SETL_DATE).isoformat()[:10]
        assert expected in exc["exception_date"]

    def test_all_exceptions_have_required_fields(self):
        """Smoke: every exception in a multi-type result has payment_id + exception_date."""
        t0 = self.PAY_DATE
        p1 = _payment_row(payment_id="pay_A", order_id="ord_A", payment_date=t0)
        p2 = _payment_row(payment_id="pay_B", order_id="ord_B", amount=500.0,
                          payment_date=t0, utr_number="UTR_B")
        p3 = _payment_row(payment_id="pay_C", order_id="ord_C",
                          payment_date=t0, utr_number="UTR_C")
        s2 = _settlement_row(payment_id="pay_B", settled_amount=480.0,
                             settlement_date=t0 + timedelta(days=1), utr_number="UTR_B")
        s3 = _settlement_row(payment_id="pay_C",
                             settlement_date=t0 + timedelta(days=5), utr_number="UTR_C")
        b2 = _bank_row(utr_number="UTR_B", bank_credited_amount=480.0)
        b3 = _bank_row(utr_number="UTR_C")
        pay_df  = pd.DataFrame([p1, p2, p3])
        pay_df["payment_date"] = pd.to_datetime(pay_df["payment_date"])
        setl_df = pd.DataFrame([s2, s3])
        setl_df["settlement_date"] = pd.to_datetime(setl_df["settlement_date"])
        bank_df = pd.DataFrame([b2, b3])[
            ["utr_number", "bank_credited_amount", "bank_credit_date", "bank_narration"]
        ]
        merged = merge_all(pay_df, setl_df, bank_df)
        result = build_result(classify(merged, sla_days=2), sla_days=2)
        for exc in result["exceptions"]:
            assert exc["payment_id"]     is not None, f"payment_id null for {exc['type']}"
            assert exc["exception_date"] is not None, f"exception_date null for {exc['type']}"


# ─────────────────────────────────────────────────────────────────────────────
# Exception structured fields — payment_id, order_id, exception_date
# ─────────────────────────────────────────────────────────────────────────────

class TestExceptionStructuredFields:
    """
    Every exception dict must expose payment_id, order_id, and exception_date
    as first-class structured fields.

    exception_date per type:
        UNMATCHED_NO_SETTLEMENT  -> payment_date
        DUPLICATE                -> payment_date
        AMOUNT_MISMATCH          -> settlement_date
        UNMATCHED_NO_BANK_CREDIT -> settlement_date
        DELAYED_SETTLEMENT       -> settlement_date
    """
    PAY_DATE  = _NOW
    SETL_DATE = _NOW + timedelta(days=3)

    def _first_exc(self, exc_type, payment=None, settlement=None, bank=None):
        pay  = payment or _payment_row(payment_date=self.PAY_DATE)
        df   = _merge_one(pay, settlement, bank)
        excs = [e for e in build_result(classify(df), sla_days=2)["exceptions"]
                if e["type"] == exc_type]
        assert excs, f"No exception of type {exc_type} generated"
        return excs[0]

    def test_unmatched_no_settlement_uses_payment_date(self):
        exc = self._first_exc(ExcType.UNMATCHED_NO_SETTLE)
        assert exc["payment_id"]     == "pay_001"
        assert exc["order_id"]       == "ord_001"
        assert exc["exception_date"] is not None
        assert "2024-01-10" in exc["exception_date"]

    def test_duplicate_uses_payment_date(self):
        t0 = self.PAY_DATE
        p1 = _payment_row(payment_id="pay_001", payment_date=t0)
        p2 = _payment_row(payment_id="pay_002", payment_date=t0 + timedelta(seconds=30))
        pay_df  = pd.DataFrame([p1, p2])
        pay_df["payment_date"] = pd.to_datetime(pay_df["payment_date"])
        setl_df = pd.DataFrame(columns=["settlement_id","payment_id","settled_amount","settlement_date","utr_number"])
        bank_df = pd.DataFrame(columns=["utr_number","bank_credited_amount","bank_credit_date","bank_narration"])
        merged  = merge_all(pay_df, setl_df, bank_df)
        result  = build_result(classify(merged, dup_window_secs=300), sla_days=2)
        dup_excs = [e for e in result["exceptions"] if e["type"] == ExcType.DUPLICATE]
        assert len(dup_excs) == 2
        for exc in dup_excs:
            assert exc["payment_id"]     is not None
            assert exc["order_id"]       == "ord_001"
            assert exc["exception_date"] is not None
            assert "2024-01-10" in exc["exception_date"]

    def test_amount_mismatch_uses_settlement_date(self):
        pay  = _payment_row(amount=1000.0, payment_date=self.PAY_DATE)
        setl = _settlement_row(settled_amount=980.0, settlement_date=self.SETL_DATE)
        exc  = self._first_exc(ExcType.AMOUNT_MISMATCH, pay, setl, _bank_row())
        assert exc["payment_id"] == "pay_001"
        assert exc["order_id"]   == "ord_001"
        expected = pd.Timestamp(self.SETL_DATE).isoformat()[:10]
        assert expected in exc["exception_date"]

    def test_no_bank_credit_uses_settlement_date(self):
        pay  = _payment_row(payment_date=self.PAY_DATE)
        setl = _settlement_row(settlement_date=self.SETL_DATE)
        exc  = self._first_exc(ExcType.UNMATCHED_NO_BANK, pay, setl, bank=None)
        assert exc["payment_id"] == "pay_001"
        assert exc["order_id"]   == "ord_001"
        expected = pd.Timestamp(self.SETL_DATE).isoformat()[:10]
        assert expected in exc["exception_date"]

    def test_delayed_settlement_uses_settlement_date(self):
        pay  = _payment_row(payment_date=self.PAY_DATE)
        setl = _settlement_row(settlement_date=self.SETL_DATE)
        exc  = self._first_exc(ExcType.DELAYED_SETTLEMENT, pay, setl, _bank_row())
        assert exc["payment_id"] == "pay_001"
        assert exc["order_id"]   == "ord_001"
        expected = pd.Timestamp(self.SETL_DATE).isoformat()[:10]
        assert expected in exc["exception_date"]

    def test_all_exceptions_have_required_fields(self):
        """Smoke: every exception in a multi-type result has payment_id + exception_date."""
        t0 = self.PAY_DATE
        p1 = _payment_row(payment_id="pay_A", order_id="ord_A", payment_date=t0)
        p2 = _payment_row(payment_id="pay_B", order_id="ord_B", amount=500.0,
                          payment_date=t0, utr_number="UTR_B")
        p3 = _payment_row(payment_id="pay_C", order_id="ord_C",
                          payment_date=t0, utr_number="UTR_C")
        s2 = _settlement_row(payment_id="pay_B", settled_amount=480.0,
                             settlement_date=t0 + timedelta(days=1), utr_number="UTR_B")
        s3 = _settlement_row(payment_id="pay_C",
                             settlement_date=t0 + timedelta(days=5), utr_number="UTR_C")
        b2 = _bank_row(utr_number="UTR_B", bank_credited_amount=480.0)
        b3 = _bank_row(utr_number="UTR_C")
        pay_df  = pd.DataFrame([p1, p2, p3])
        pay_df["payment_date"] = pd.to_datetime(pay_df["payment_date"])
        setl_df = pd.DataFrame([s2, s3])
        setl_df["settlement_date"] = pd.to_datetime(setl_df["settlement_date"])
        bank_df = pd.DataFrame([b2, b3])[
            ["utr_number", "bank_credited_amount", "bank_credit_date", "bank_narration"]
        ]
        merged = merge_all(pay_df, setl_df, bank_df)
        result = build_result(classify(merged, sla_days=2), sla_days=2)
        for exc in result["exceptions"]:
            assert exc["payment_id"]     is not None, f"payment_id null for {exc['type']}"
            assert exc["exception_date"] is not None, f"exception_date null for {exc['type']}"

# ─────────────────────────────────────────────────────────────────────────────
# API Tests
# ─────────────────────────────────────────────────────────────────────────────

from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from app.main import app
from app.database import Base, get_db
from app.models import ReconciliationRun, ReconciliationStatus

engine = create_engine(
    "sqlite:///:memory:",
    connect_args={"check_same_thread": False},
    poolclass=StaticPool
)
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

def override_get_db():
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()

app.dependency_overrides[get_db] = override_get_db
client = TestClient(app)

def test_get_runs_list():
    Base.metadata.create_all(bind=engine)
    db = TestingSessionLocal()
    try:
        # 1. Insert 3 runs with different started_at dates
        # r3 is oldest, r1 is next, r2 is newest
        now = datetime.now(timezone.utc)
        r1 = ReconciliationRun(status=ReconciliationStatus.COMPLETED, started_at=now - timedelta(days=2))
        r2 = ReconciliationRun(status=ReconciliationStatus.COMPLETED, started_at=now - timedelta(days=1))
        r3 = ReconciliationRun(status=ReconciliationStatus.COMPLETED, started_at=now - timedelta(days=3))
        db.add_all([r1, r2, r3])
        db.commit()

        # 2. test descending order
        resp = client.get("/api/v1/reconcile/runs")
        assert resp.status_code == 200
        data = resp.json()
        assert data["total"] == 3
        runs = data["runs"]
        assert len(runs) == 3
        # r2 newest -> r1 -> r3 oldest
        assert runs[0]["run_id"] == r2.id
        assert runs[1]["run_id"] == r1.id
        assert runs[2]["run_id"] == r3.id

        # 3. test pagination parameter
        resp_skip = client.get("/api/v1/reconcile/runs?skip=1&limit=1")
        assert resp_skip.status_code == 200
        data_skip = resp_skip.json()
        assert len(data_skip["runs"]) == 1
        assert data_skip["runs"][0]["run_id"] == r1.id
    finally:
        db.query(ReconciliationRun).delete()
        db.commit()
        db.close()


def test_get_summary_trends():
    """
    Test that /api/v1/summary/trends calculates match rate exactly as
    /api/v1/reconcile/summary/{run_id} expects or computes.
    """
    Base.metadata.create_all(bind=engine)
    db = TestingSessionLocal()
    try:
        now = datetime.now(timezone.utc)
        # total=10, matched=6 => match_rate = 60.0
        r1 = ReconciliationRun(
            status=ReconciliationStatus.COMPLETED,
            started_at=now - timedelta(days=1),
            total_transactions=10,
            matched_count=6,
            unmatched_count=4,
            exception_count=0
        )
        # total=0 => match_rate = 0.0 (edge case)
        r2 = ReconciliationRun(
            status=ReconciliationStatus.COMPLETED,
            started_at=now,
            total_transactions=0,
            matched_count=0,
            unmatched_count=0,
            exception_count=0
        )
        db.add_all([r1, r2])
        db.commit()
        db.refresh(r1)
        db.refresh(r2)

        resp = client.get("/api/v1/summary/trends?days=2")
        assert resp.status_code == 200, resp.text
        data = resp.json()["data_points"]
        
        # We expect 2 points
        assert len(data) >= 2
        
        rate_r1 = next(d["match_rate"] for d in data if d["run_id"] == r1.id)
        rate_r2 = next(d["match_rate"] for d in data if d["run_id"] == r2.id)

        assert rate_r1 == 60.0
        assert rate_r2 == 0.0
    finally:
        db.query(ReconciliationRun).delete()
        db.commit()
        db.close()


def test_delete_single_run():
    Base.metadata.create_all(bind=engine)
    db = TestingSessionLocal()
    try:
        from app.models import Exception as ExceptionModel
        run = ReconciliationRun(status="COMPLETED", total_transactions=5)
        db.add(run)
        db.commit()
        db.refresh(run)

        exc = ExceptionModel(reconciliation_run_id=run.id, exception_type="test", severity="medium")
        db.add(exc)
        db.commit()

        # Delete it using the API
        resp = client.delete(f"/api/v1/reconcile/runs/{run.id}")
        assert resp.status_code == 200
        assert resp.json()["run_id"] == run.id

        # Verify it was completely removed from the database
        assert db.query(ReconciliationRun).filter_by(id=run.id).count() == 0
        assert db.query(ExceptionModel).filter_by(reconciliation_run_id=run.id).count() == 0

        # Try deleting non-existent
        resp2 = client.delete("/api/v1/reconcile/runs/99999")
        assert resp2.status_code == 404
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────────────────────
# Chat Assistant Tests
# ─────────────────────────────────────────────────────────────────────────────

class TestChatAssistant:
    """
    Test chat feature: GET_CHAT_RESPONSE function and POST /api/v1/reconcile/chat endpoint
    """

    def test_chat_with_valid_run_returns_response(self):
        """Test that chat endpoint returns a response for a valid run_id"""
        from app.main import app
        from fastapi.testclient import TestClient
        from unittest.mock import patch

        client = TestClient(app)

        Base.metadata.create_all(bind=engine)
        db = TestingSessionLocal()
        try:
            from app.models import ReconciliationRun, Exception as ExceptionModel

            run = ReconciliationRun(
                status=ReconciliationStatus.COMPLETED,
                run_name="Test Run for Chat",
                total_transactions=10,
                matched_count=8,
                unmatched_count=2,
                exception_count=2,
            )
            db.add(run)
            db.commit()
            db.refresh(run)

            # Add a test exception
            exc = ExceptionModel(
                reconciliation_run_id=run.id,
                exception_type="AMOUNT_MISMATCH",
                severity="medium",
                payment_id="pay_001",
                order_id="ord_001",
                internal_amount=100.0,
                bank_amount=90.0,
                discrepancy_amount=10.0,
                is_resolved=False,
                description="Test amount mismatch",
            )
            db.add(exc)
            db.commit()

            with patch("app.services.chat_assistant.get_gemini_client", return_value=object()):
                with patch(
                    "app.services.chat_assistant.generate_gemini_content",
                    return_value=("This run has 2 exceptions.", "gemini-3.5-flash"),
                ):
                    resp = client.post(
                        "/api/v1/reconcile/chat",
                        json={
                            "run_id": run.id,
                            "message": "How many exceptions are in this run?",
                            "conversation_history": [],
                        },
                    )

            assert resp.status_code == 200
            data = resp.json()
            assert data["response"] == "This run has 2 exceptions."

        finally:
            from app.models import Exception as ExceptionModel
            if "run" in locals():
                db.query(ExceptionModel).filter(ExceptionModel.reconciliation_run_id == run.id).delete()
                db.query(ReconciliationRun).filter(ReconciliationRun.id == run.id).delete()
                db.commit()
            db.close()

    def test_chat_with_invalid_run_returns_404(self):
        """Test that chat endpoint returns 404 for non-existent run"""
        from app.main import app
        from fastapi.testclient import TestClient

        client = TestClient(app)

        resp = client.post(
            "/api/v1/reconcile/chat",
            json={
                "run_id": 99999,
                "message": "What is this run?",
                "conversation_history": [],
            },
        )

        assert resp.status_code == 404
        data = resp.json()
        assert "not found" in data["detail"].lower()

    def test_chat_system_prompt_includes_real_data(self):
        """Test that system prompt is constructed with actual run data and exceptions"""
        from app.services.chat_assistant import get_chat_response
        from unittest.mock import patch, MagicMock

        Base.metadata.create_all(bind=engine)
        db = TestingSessionLocal()
        try:
            from app.models import ReconciliationRun, Exception as ExceptionModel

            run = ReconciliationRun(
                status=ReconciliationStatus.COMPLETED,
                run_name="Data Test Run",
                total_transactions=5,
                matched_count=4,
                unmatched_count=1,
                exception_count=1,
            )
            db.add(run)
            db.commit()
            db.refresh(run)

            # Add exception with specific payment_id and order_id
            exc = ExceptionModel(
                reconciliation_run_id=run.id,
                exception_type="DUPLICATE",
                severity="high",
                payment_id="pay_test_12345",
                order_id="ord_test_99999",
                internal_amount=500.0,
                bank_amount=0.0,
                discrepancy_amount=500.0,
                is_resolved=False,
                description="Test duplicate payment",
            )
            db.add(exc)
            db.commit()

            captured_prompt = {}

            def capture_generate_content(client, contents):
                captured_prompt["contents"] = contents
                return (
                    "This run has 1 duplicate exception for payment pay_test_12345.",
                    "gemini-3.5-flash",
                )

            with patch("app.services.chat_assistant.get_gemini_client", return_value=MagicMock()):
                with patch(
                    "app.services.chat_assistant.generate_gemini_content",
                    side_effect=capture_generate_content,
                ):
                    response = get_chat_response(
                        run_id=run.id,
                        message="Tell me about the exceptions.",
                        conversation_history=[],
                        db=db,
                    )

            # Verify that response was generated
            assert response is not None
            assert len(response) > 0

            # Verify that system prompt includes actual data from the run
            contents = captured_prompt.get("contents", [])
            prompt_text = contents[0].parts[0].text
            
            # Check for run name
            assert "Data Test Run" in prompt_text, "System prompt missing run name"
            
            # Check for payment_id from the exception
            assert "pay_test_12345" in prompt_text, "System prompt missing payment_id"
            
            # Check for order_id from the exception
            assert "ord_test_99999" in prompt_text, "System prompt missing order_id"
            
            # Check for exception type
            assert "DUPLICATE" in prompt_text, "System prompt missing exception type"

            assert "matched_count" in prompt_text, "System prompt missing run summary"
            assert "discrepancy_amount" in prompt_text, "System prompt missing exception amounts"
            assert "using ONLY this data" in prompt_text, "System prompt missing data-only constraint"

        finally:
            from app.models import Exception as ExceptionModel
            if "run" in locals():
                db.query(ExceptionModel).filter(ExceptionModel.reconciliation_run_id == run.id).delete()
                db.query(ReconciliationRun).filter(ReconciliationRun.id == run.id).delete()
                db.commit()
            db.close()

    def test_chat_with_conversation_history_preserves_context(self):
        """Test that multi-turn conversation history is passed to Gemini"""
        from app.services.chat_assistant import get_chat_response
        from unittest.mock import patch, MagicMock

        Base.metadata.create_all(bind=engine)
        db = TestingSessionLocal()
        try:
            from app.models import ReconciliationRun, Exception as ExceptionModel

            run = ReconciliationRun(
                status=ReconciliationStatus.COMPLETED,
                run_name="Conversation Test Run",
                total_transactions=3,
                matched_count=3,
                unmatched_count=0,
                exception_count=0,
            )
            db.add(run)
            db.commit()
            db.refresh(run)

            # Mock Gemini to capture conversation contents
            captured_contents = {}

            def capture_generate_content(client, contents):
                captured_contents["contents"] = contents
                return "Yes, that's correct based on the data.", "gemini-3.5-flash"

            with patch("app.services.chat_assistant.get_gemini_client", return_value=MagicMock()):
                with patch(
                    "app.services.chat_assistant.generate_gemini_content",
                    side_effect=capture_generate_content,
                ):
                    # Include conversation history
                    response = get_chat_response(
                        run_id=run.id,
                        message="Is the match rate 100%?",
                        conversation_history=[
                            {"role": "user", "content": "What is the match rate?"},
                            {
                                "role": "assistant",
                                "content": "The match rate is 100% (3/3 matched)",
                            },
                        ],
                        db=db,
                    )

            # Verify response
            assert response is not None

            # Verify conversation history was passed
            contents = captured_contents.get("contents", [])
            contents_str = str(contents)
            
            # Check that prior conversation is in the contents
            assert "What is the match rate?" in contents_str, "Prior user message missing"
            assert "100%" in contents_str, "Prior assistant message missing"

        finally:
            from app.models import Exception as ExceptionModel
            if "run" in locals():
                db.query(ExceptionModel).filter(ExceptionModel.reconciliation_run_id == run.id).delete()
                db.query(ReconciliationRun).filter(ReconciliationRun.id == run.id).delete()
                db.commit()
            db.close()


def test_delete_all_runs():
    Base.metadata.create_all(bind=engine)
    db = TestingSessionLocal()
    try:
        from app.models import Exception as ExceptionModel
        db.add(ReconciliationRun(status="COMPLETED", total_transactions=5))
        db.add(ReconciliationRun(status="COMPLETED", total_transactions=2))
        db.add(ReconciliationRun(status="COMPLETED", total_transactions=8))
        db.commit()

        # Without confirm, should fail 400
        resp1 = client.delete("/api/v1/reconcile/runs")
        assert resp1.status_code == 400
        assert "confirm=true" in resp1.json()["detail"].lower()

        # With confirm, should succeed 200 and clear the database returning accurate count
        resp2 = client.delete("/api/v1/reconcile/runs?confirm=true")
        assert resp2.status_code == 200
        assert "deleted" in resp2.json()["message"].lower()
        assert resp2.json()["count"] == 3

        assert db.query(ReconciliationRun).count() == 0
    finally:
        from app.models import Exception as ExceptionModel
        db.query(ExceptionModel).delete()
        db.query(ReconciliationRun).delete()
        db.commit()
        db.close()


def test_orphan_bank_credit_creates_exception():
    """A bank statement row with no matching settlement produces 1 UNMATCHED_NO_BANK_CREDIT exception."""
    import pandas as pd
    from app.services.reconciliation import find_orphan_bank_credits, ExcType

    settlements = pd.DataFrame([
        {"settlement_id": "setl_1", "payment_id": "pay_1", "settled_amount": 100.0, "utr_number": "UTR111"}
    ])
    bank = pd.DataFrame([
        {"utr_number": "UTR111", "credited_amount": 100.0, "credit_date": "2024-06-01", "narration": "NORMAL CREDIT"},
        {"utr_number": "UTR999", "credited_amount": 500.0, "credit_date": "2024-06-02", "narration": "GHOST CREDIT"}
    ])

    orphans = find_orphan_bank_credits(settlements, bank)
    assert len(orphans) == 1
    assert orphans[0]["type"] == ExcType.UNMATCHED_NO_BANK
    assert orphans[0]["utr_number"] == "UTR999"
    assert orphans[0]["amount"] == 500.0
    assert orphans[0]["severity"] == "high"
    assert orphans[0]["payment_id"] is None
    assert orphans[0]["order_id"] is None


def test_matched_bank_credit_creates_no_orphan():
    """A bank statement row that DOES match a settlement produces no orphan exception."""
    import pandas as pd
    from app.services.reconciliation import find_orphan_bank_credits

    settlements = pd.DataFrame([
        {"settlement_id": "setl_1", "payment_id": "pay_1", "settled_amount": 100.0, "utr_number": "UTR111"}
    ])
    bank = pd.DataFrame([
        {"utr_number": "UTR111", "credited_amount": 100.0, "credit_date": "2024-06-01", "narration": "MATCHED"}
    ])

    orphans = find_orphan_bank_credits(settlements, bank)
    assert len(orphans) == 0


def test_orphan_bank_credit_preserves_payment_invariants():
    """Adding orphan bank rows does not change total_records, matched, or match_rate."""
    import pandas as pd
    from app.services.reconciliation import classify, build_result

    payments = pd.DataFrame([
        {"payment_id": "pay_1", "order_id": "ord_1", "amount": 100.0, "status": "captured", "payment_date": "2024-06-01"},
        {"payment_id": "pay_2", "order_id": "ord_2", "amount": 200.0, "status": "captured", "payment_date": "2024-06-01"}
    ])
    settlements = pd.DataFrame([
        {"settlement_id": "setl_1", "payment_id": "pay_1", "settled_amount": 100.0, "utr_number": "UTR111"}
    ])
    bank = pd.DataFrame([
        {"utr_number": "UTR111", "credited_amount": 100.0, "credit_date": "2024-06-01", "narration": "MATCHED"},
        {"utr_number": "UTR888", "credited_amount": 300.0, "credit_date": "2024-06-02", "narration": "GHOST 1"},
        {"utr_number": "UTR777", "credited_amount": 400.0, "credit_date": "2024-06-02", "narration": "GHOST 2"}
    ])

    merged = payments.merge(settlements, on="payment_id", how="left").merge(bank, on="utr_number", how="left")
    classified = classify(merged)
    res = build_result(classified, sla_days=2, settlements=settlements, bank=bank)

    # Invariants
    assert res["total_records"] == 2
    assert res["matched"] == 1
    assert res["match_rate"] == 50.0
    assert res["unmatched_bank_credits"]["count"] == 2
    assert res["unmatched_bank_credits"]["total_amount"] == 700.0
    # Exceptions total = 1 payment unmatched + 2 orphan bank credits = 3
    assert len(res["exceptions"]) == 3


def test_duplicate_bank_credit_creates_exception():
    """Duplicate bank statement rows sharing the same UTR produce a DUPLICATE_BANK_CREDIT exception."""
    import pandas as pd
    from app.services.reconciliation import find_duplicate_bank_credits, ExcType

    bank = pd.DataFrame([
        {"utr_number": "UTR111", "credited_amount": 100.0, "credit_date": "2024-06-01", "narration": "CREDIT 1"},
        {"utr_number": "UTR111", "credited_amount": 100.0, "credit_date": "2024-06-01", "narration": "DUPLICATE CREDIT"}
    ])

    dups = find_duplicate_bank_credits(bank)
    assert len(dups) == 1
    assert dups[0]["type"] == ExcType.DUPLICATE_BANK
    assert dups[0]["utr_number"] == "UTR111"
    assert dups[0]["amount"] == 100.0
    assert dups[0]["severity"] == "high"


def test_ai_analysis_caching_and_fallback():
    """Test AI analysis caching and DB persistence behavior."""
    # Use the module-level in-memory test DB (TestingSessionLocal/engine), NOT the
    # production app.database.SessionLocal — importing the real session here leaked
    # permanent "Test AI Run" rows into backend/reconciliation.db.
    from app.models import Exception as ExceptionModel
    from app.services.ai_analysis import analyze_exceptions_for_run

    Base.metadata.create_all(bind=engine)
    db = TestingSessionLocal()
    run = None
    exc = None
    try:
        run = ReconciliationRun(run_name="Test AI Run", status=ReconciliationStatus.COMPLETED)
        db.add(run)
        db.commit()
        db.refresh(run)

        exc = ExceptionModel(
            reconciliation_run_id=run.id,
            exception_type="AMOUNT_MISMATCH",
            severity="medium",
            payment_id="pay_test_ai",
            description="Amount mismatch test",
            internal_amount=100.0,
            bank_amount=90.0,
            discrepancy_amount=10.0,
        )
        db.add(exc)
        db.commit()

        # First call will generate or use fallback
        res1 = analyze_exceptions_for_run(db, run.id)
        assert len(res1) == 1
        assert res1[0]["id"] == exc.id

        # Update DB row explicitly with cached text
        exc.ai_explanation = "CACHED_EXPLANATION_TEST"
        db.commit()

        # Second call must serve cached value directly from DB without calling Gemini
        res2 = analyze_exceptions_for_run(db, run.id)
        assert len(res2) == 1
        assert res2[0]["explanation"] == "CACHED_EXPLANATION_TEST"
    finally:
        # Clean up the rows this test created so nothing accumulates, even if an
        # assertion above fails mid-test.
        try:
            if run is not None:
                db.query(ExceptionModel).filter(
                    ExceptionModel.reconciliation_run_id == run.id
                ).delete(synchronize_session=False)
                db.query(ReconciliationRun).filter(
                    ReconciliationRun.id == run.id
                ).delete(synchronize_session=False)
                db.commit()
        finally:
            db.close()


# ─────────────────────────────────────────────────────────────────────────────
# Excel Report Download Tests
# ─────────────────────────────────────────────────────────────────────────────

from app.models import Exception as ExceptionModel

def test_download_report_valid_run():
    """GET /report/{run_id} returns a valid non-empty .xlsx for a real run."""
    Base.metadata.create_all(bind=engine)
    db = TestingSessionLocal()
    try:
        # Create a run with realistic data
        run = ReconciliationRun(
            run_name="Report Test Run",
            status=ReconciliationStatus.COMPLETED,
            total_transactions=10,
            matched_count=7,
            unmatched_count=3,
            exception_count=3,
            ai_insights='{"match_rate": 70.0, "settlement_totals": {"total_payment_amount": 10000, "total_settled_amount": 8000, "total_discrepancy": 2000, "amount_mismatches_total": 500, "unsettled_value_total": 1200, "duplicate_charges_total": 300, "currency": "INR"}, "exception_counts": {"AMOUNT_MISMATCH": 1, "UNMATCHED_NO_SETTLEMENT": 1, "DUPLICATE": 1}}',
        )
        db.add(run)
        db.commit()
        db.refresh(run)

        exc1 = ExceptionModel(
            reconciliation_run_id=run.id,
            exception_type="AMOUNT_MISMATCH",
            severity="medium",
            payment_id="pay_report_test_1",
            description="Amount mismatch test",
            internal_amount=100.0,
            bank_amount=90.0,
            discrepancy_amount=10.0,
            ai_explanation="Payment discrepancy identified due to gateway fee deduction.",
        )
        db.add(exc1)
        db.commit()

        resp = client.get(f"/api/v1/reconcile/report/{run.id}")
        assert resp.status_code == 200, f"Expected 200, got {resp.status_code}: {resp.text}"
        assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment" in resp.headers.get("content-disposition", "")
        assert f"run_{run.id}" in resp.headers.get("content-disposition", "")

        # Verify the response body is a valid, non-empty .xlsx
        content = resp.content
        assert len(content) > 100, "Response body too small to be a valid .xlsx"
        # .xlsx files are ZIP archives that start with PK signature
        assert content[:2] == b"PK", "Response does not start with PK (ZIP/xlsx) signature"

        # Verify we can actually open it with openpyxl
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content))
        assert "Summary" in wb.sheetnames
        assert "Exceptions" in wb.sheetnames
        assert "Raw Reconciliation" in wb.sheetnames

        # Verify AI Explanation column in Exceptions sheet is populated from DB cache
        ws_exc = wb["Exceptions"]
        assert ws_exc.cell(row=1, column=11).value == "AI Explanation"
        assert ws_exc.cell(row=2, column=11).value == "Payment discrepancy identified due to gateway fee deduction."
        wb.close()
    finally:
        try:
            if run is not None:
                db.query(ExceptionModel).filter(
                    ExceptionModel.reconciliation_run_id == run.id
                ).delete(synchronize_session=False)
                db.query(ReconciliationRun).filter(
                    ReconciliationRun.id == run.id
                ).delete(synchronize_session=False)
                db.commit()
        finally:
            db.close()


def test_download_report_ai_explanation_from_cache():
    """Verify generate_excel_report populates AI Explanation purely from cached DB rows without calling Gemini."""
    Base.metadata.create_all(bind=engine)
    db = TestingSessionLocal()
    run = None
    try:
        run = ReconciliationRun(
            run_name="Cache Only Report Test",
            status=ReconciliationStatus.COMPLETED,
            total_transactions=2,
            matched_count=1,
            unmatched_count=1,
            exception_count=1,
            ai_insights='{"match_rate": 50.0, "settlement_totals": {}, "exception_counts": {}}',
        )
        db.add(run)
        db.commit()
        db.refresh(run)

        exc = ExceptionModel(
            reconciliation_run_id=run.id,
            exception_type="UNMATCHED_NO_SETTLEMENT",
            severity="high",
            payment_id="pay_cache_test_99",
            description="Unsettled payment",
            internal_amount=500.0,
            ai_explanation="Customer paid via UPI but settlement batch was delayed past SLA window.",
        )
        db.add(exc)
        db.commit()

        # Generate report
        from app.services.report_generator import generate_excel_report
        stream = generate_excel_report(run.id, db)
        assert stream is not None

        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(stream.getvalue()))
        ws = wb["Exceptions"]
        # Row 2 (first exception row), Column 11 (AI Explanation)
        assert ws.cell(row=2, column=11).value == "Customer paid via UPI but settlement batch was delayed past SLA window."
        wb.close()
    finally:
        try:
            if run is not None:
                db.query(ExceptionModel).filter(
                    ExceptionModel.reconciliation_run_id == run.id
                ).delete(synchronize_session=False)
                db.query(ReconciliationRun).filter(
                    ReconciliationRun.id == run.id
                ).delete(synchronize_session=False)
                db.commit()
        finally:
            db.close()


def test_download_report_invalid_run():
    """GET /report/{run_id} returns 404 for a non-existent run."""
    resp = client.get("/api/v1/reconcile/report/999999")
    assert resp.status_code == 404


# ─────────────────────────────────────────────────────────────────────────────
# Upload Validation Hardening Tests
# ─────────────────────────────────────────────────────────────────────────────

_VALID_PAYMENTS = b"payment_id,order_id,amount,currency,payment_date,method,status,utr_number\npay_1,ord_1,100.0,INR,2024-01-01,upi,captured,UTR1\n"
_VALID_SETTLEMENTS = b"settlement_id,payment_id,settled_amount,settlement_date,utr_number\nsetl_1,pay_1,100.0,2024-01-02,UTR1\n"
_VALID_BANK = b"utr_number,credited_amount,credit_date,narration\nUTR1,100.0,2024-01-02,CREDIT\n"


def test_upload_empty_csv_rejected():
    """Empty CSV (0 bytes or headers only) returns 400 with file name."""
    # Case A: 0 bytes
    resp = client.post(
        "/api/v1/reconcile/upload",
        files={
            "payments": ("payments.csv", b"", "text/csv"),
            "settlements": ("settlements.csv", _VALID_SETTLEMENTS, "text/csv"),
            "bank_statement": ("bank_statement.csv", _VALID_BANK, "text/csv"),
        },
    )
    assert resp.status_code == 400
    detail = resp.json()["detail"]
    assert "payments.csv" in detail
    assert "empty" in detail.lower()

    # Case B: Headers only, no data rows
    headers_only = b"payment_id,order_id,amount,currency,payment_date,method,status,utr_number\n"
    resp2 = client.post(
        "/api/v1/reconcile/upload",
        files={
            "payments": ("payments.csv", headers_only, "text/csv"),
            "settlements": ("settlements.csv", _VALID_SETTLEMENTS, "text/csv"),
            "bank_statement": ("bank_statement.csv", _VALID_BANK, "text/csv"),
        },
    )
    assert resp2.status_code == 400
    detail2 = resp2.json()["detail"]
    assert "payments.csv" in detail2
    assert "no data rows" in detail2.lower()


def test_upload_missing_required_columns_rejected():
    """Missing required columns returns 400 naming missing columns and file."""
    bad_settlements = b"settlement_id,payment_id,settlement_date\nsetl_1,pay_1,2024-01-02\n"
    resp = client.post(
        "/api/v1/reconcile/upload",
        files={
            "payments": ("payments.csv", _VALID_PAYMENTS, "text/csv"),
            "settlements": ("settlements.csv", bad_settlements, "text/csv"),
            "bank_statement": ("bank_statement.csv", _VALID_BANK, "text/csv"),
        },
    )
    assert resp.status_code == 400
    detail = resp.json()["detail"]
    assert "settlements.csv" in detail
    assert "missing required columns" in detail.lower()
    assert "settled_amount" in detail
    assert "utr_number" in detail


def test_upload_non_numeric_amount_rejected():
    """Non-numeric value in amount column returns 400 with file, column, and row number."""
    bad_payments = (
        b"payment_id,order_id,amount,currency,payment_date,method,status,utr_number\n"
        b"pay_1,ord_1,100.0,INR,2024-01-01,upi,captured,UTR1\n"
        b"pay_2,ord_2,INVALID_AMT,INR,2024-01-01,upi,captured,UTR2\n"
    )
    resp = client.post(
        "/api/v1/reconcile/upload",
        files={
            "payments": ("payments.csv", bad_payments, "text/csv"),
            "settlements": ("settlements.csv", _VALID_SETTLEMENTS, "text/csv"),
            "bank_statement": ("bank_statement.csv", _VALID_BANK, "text/csv"),
        },
    )
    assert resp.status_code == 400
    detail = resp.json()["detail"]
    assert "payments.csv" in detail
    assert "amount" in detail.lower()
    assert "row 3" in detail.lower()
    assert "INVALID_AMT" in detail


def test_upload_malformed_csv_rejected():
    """Malformed CSV content returns 400 with parse error."""
    # Broken quotes / incompatible structure
    malformed_bank = b'utr_number,credited_amount,credit_date\n"UTR1",100.0,"2024-01-02\nUTR2,200.0,2024-01-03\n'
    resp = client.post(
        "/api/v1/reconcile/upload",
        files={
            "payments": ("payments.csv", _VALID_PAYMENTS, "text/csv"),
            "settlements": ("settlements.csv", _VALID_SETTLEMENTS, "text/csv"),
            "bank_statement": ("bank_statement.csv", malformed_bank, "text/csv"),
        },
    )
    assert resp.status_code == 400
    detail = resp.json()["detail"]
    assert "bank_statement.csv" in detail


def test_upload_wrong_file_extension_rejected():
    """Non-CSV file extension returns 400 with clear message."""
    resp = client.post(
        "/api/v1/reconcile/upload",
        files={
            "payments": ("payments.xlsx", _VALID_PAYMENTS, "application/octet-stream"),
            "settlements": ("settlements.csv", _VALID_SETTLEMENTS, "text/csv"),
            "bank_statement": ("bank_statement.csv", _VALID_BANK, "text/csv"),
        },
    )
    assert resp.status_code == 400
    detail = resp.json()["detail"]
    assert "payments.xlsx" in detail
    assert "not a CSV file" in detail


def test_raw_reconciliation_sheet_exact_63_rows_for_run_1():
    """
    Permanent regression guard: Raw Reconciliation sheet for a standard sample run
    must contain EXACTLY 63 data rows (43 matched + 20 payment-level exceptions),
    excluding header. Fails if join fan-out regression occurs. Self-contained fixture.
    """
    import io
    import json
    from pathlib import Path
    from openpyxl import load_workbook
    from app.services.report_generator import generate_excel_report
    from app.services.reconciliation import ReconciliationEngine

    Base.metadata.create_all(bind=engine)
    db = TestingSessionLocal()
    run = None
    try:
        data_dir = Path("sample_data")
        if not data_dir.exists():
            data_dir = Path("../sample_data")

        engine_inst = ReconciliationEngine(
            payments_path=data_dir / "payments.csv",
            settlements_path=data_dir / "settlements.csv",
            bank_statement_path=data_dir / "bank_statement.csv",
            sla_days=2,
        )
        result = engine_inst.run()

        run = ReconciliationRun(
            run_name="Self-contained Sample Run",
            status=ReconciliationStatus.COMPLETED,
            total_transactions=result["total_records"],
            matched_count=result["matched"],
            unmatched_count=result["unmatched"],
            exception_count=len(result.get("exceptions", [])),
            ai_insights=json.dumps({
                "settlement_totals": result.get("settlement_totals", {}),
                "exception_counts": result.get("exception_counts", {}),
                "match_rate": result["match_rate"],
            }),
        )
        db.add(run)
        db.commit()
        db.refresh(run)

        # Populate exceptions
        for e in result.get("exceptions", []):
            db.add(
                ExceptionModel(
                    reconciliation_run_id=run.id,
                    exception_type=e.get("type"),
                    severity=e.get("severity", "medium"),
                    payment_id=e.get("payment_id"),
                    order_id=e.get("order_id"),
                    exception_date=e.get("exception_date"),
                    description=e.get("details") or e.get("description"),
                    internal_amount=e.get("amount") or e.get("internal_amount"),
                    bank_amount=e.get("settled_amount") or e.get("bank_amount"),
                    discrepancy_amount=e.get("amount_impact") or e.get("discrepancy_amount"),
                )
            )
        db.commit()

        stream = generate_excel_report(run.id, db)
        wb = load_workbook(stream)
        assert "Raw Reconciliation" in wb.sheetnames, "Missing 'Raw Reconciliation' sheet"
        ws_raw = wb["Raw Reconciliation"]

        rows = list(ws_raw.iter_rows(values_only=True))
        assert len(rows) > 0, "Raw Reconciliation sheet is empty"

        header = rows[0]
        expected_header = (
            "payment_id",
            "order_id",
            "payment_amount",
            "settled_amount",
            "bank_credited_amount",
            "utr_number",
            "payment_date",
            "settlement_date",
            "status",
        )
        assert header == expected_header, f"Header mismatch: {header} vs {expected_header}"

        data_rows = rows[1:]
        assert len(data_rows) == 63, f"Expected exactly 63 data rows in Raw Reconciliation, found {len(data_rows)}"

        # Status breakdown validation
        from collections import Counter
        counts = Counter(r[8] for r in data_rows)
        assert counts["MATCHED"] == 43, f"Expected 43 MATCHED rows, found {counts['MATCHED']}"
        assert counts["DUPLICATE"] == 6, f"Expected 6 DUPLICATE rows, found {counts['DUPLICATE']}"
        assert counts["DELAYED_SETTLEMENT"] == 5, f"Expected 5 DELAYED_SETTLEMENT rows, found {counts['DELAYED_SETTLEMENT']}"
        assert counts["UNMATCHED_NO_SETTLEMENT"] == 5, f"Expected 5 UNMATCHED_NO_SETTLEMENT rows, found {counts['UNMATCHED_NO_SETTLEMENT']}"
        assert counts["AMOUNT_MISMATCH"] == 4, f"Expected 4 AMOUNT_MISMATCH rows, found {counts['AMOUNT_MISMATCH']}"
        assert sum(counts.values()) == 63
        wb.close()
    finally:
        if run is not None:
            db.query(ExceptionModel).filter(ExceptionModel.reconciliation_run_id == run.id).delete(synchronize_session=False)
            db.query(ReconciliationRun).filter(ReconciliationRun.id == run.id).delete(synchronize_session=False)
            db.commit()
        db.close()




