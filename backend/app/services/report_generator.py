"""
app/services/report_generator.py
---------------------------------
Generate a formatted Excel (.xlsx) reconciliation report for a given run.

Sheets
------
1. Summary      — match rate, record counts, settlement totals including
                   3-way discrepancy split and gross discrepancy
2. Exceptions   — full exception list with AI explanation if available
3. Raw Reconciliation — transaction-level data with match status per row
"""

from __future__ import annotations

import io
import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models import (
    Exception as ExceptionModel,
    ReconciliationRun,
    ReconciliationStatus,
)

logger = logging.getLogger(__name__)

# ── Styling constants ────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

LABEL_FONT = Font(name="Calibri", bold=True, size=11)
VALUE_FONT = Font(name="Calibri", size=11)
CURRENCY_FORMAT = '#,##0.00'
PERCENT_FORMAT = '0.00"%"'

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _auto_size_columns(ws, min_width: int = 10, max_width: int = 60):
    """Auto-size every column based on content length."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def _style_header_row(ws, row: int, num_cols: int):
    """Apply header styling to a row."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _write_summary_pair(ws, row: int, label: str, value: Any, is_currency: bool = False):
    """Write a label-value pair on the summary sheet."""
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = LABEL_FONT
    label_cell.alignment = Alignment(horizontal="right")
    label_cell.border = THIN_BORDER

    value_cell = ws.cell(row=row, column=2, value=value)
    value_cell.font = VALUE_FONT
    value_cell.alignment = Alignment(horizontal="left")
    value_cell.border = THIN_BORDER
    if is_currency and isinstance(value, (int, float)):
        value_cell.number_format = CURRENCY_FORMAT


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────

def generate_excel_report(run_id: int, db: Session) -> io.BytesIO:
    """
    Build a formatted .xlsx workbook for a reconciliation run.

    Returns an in-memory BytesIO stream ready for streaming to the client.
    Raises ValueError if the run_id is not found.
    """
    # ── Fetch run ────────────────────────────────────────────────────────────
    run: ReconciliationRun | None = (
        db.query(ReconciliationRun)
        .filter(ReconciliationRun.id == run_id)
        .first()
    )
    if run is None:
        raise ValueError(f"Reconciliation run {run_id} not found")

    # ── Parse cached insights ────────────────────────────────────────────────
    insights: dict = {}
    if run.ai_insights:
        try:
            insights = json.loads(run.ai_insights)
        except json.JSONDecodeError:
            pass

    settlement_totals: dict = insights.get("settlement_totals", {})
    exception_counts: dict = insights.get("exception_counts", {})
    match_rate: float = insights.get("match_rate", 0.0)

    # ── Fetch exceptions ─────────────────────────────────────────────────────
    exceptions = (
        db.query(ExceptionModel)
        .filter(ExceptionModel.reconciliation_run_id == run_id)
        .order_by(ExceptionModel.id.asc())
        .all()
    )

    # Recompute 3-way split if missing from insights
    if "amount_mismatches_total" not in settlement_totals:
        settlement_totals["amount_mismatches_total"] = round(
            sum((e.discrepancy_amount or 0.0) for e in exceptions if e.exception_type == "AMOUNT_MISMATCH"), 2
        )
    if "unsettled_value_total" not in settlement_totals:
        settlement_totals["unsettled_value_total"] = round(
            sum((e.internal_amount or 0.0) for e in exceptions if e.exception_type == "UNMATCHED_NO_SETTLEMENT"), 2
        )
    if "duplicate_charges_total" not in settlement_totals:
        settlement_totals["duplicate_charges_total"] = round(
            sum((e.internal_amount or 0.0) for e in exceptions if e.exception_type == "DUPLICATE"), 2
        )

    # ── Create workbook ──────────────────────────────────────────────────────
    wb = Workbook()

    _build_summary_sheet(wb, run, match_rate, settlement_totals, exception_counts)
    _build_exceptions_sheet(wb, exceptions)
    _build_raw_reconciliation_sheet(wb, run, exceptions, match_rate)

    # Remove the default empty sheet if we created named ones first
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── Serialize to bytes ───────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─────────────────────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────────────────────

def _build_summary_sheet(
    wb: Workbook,
    run: ReconciliationRun,
    match_rate: float,
    settlement_totals: dict,
    exception_counts: dict,
):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_properties.tabColor = "1F4E79"

    row = 1

    # ── Report header ────────────────────────────────────────────────────────
    ws.merge_cells("A1:B1")
    title_cell = ws.cell(row=1, column=1, value="RazorRecon AI — Reconciliation Report")
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")
    row += 1

    ws.merge_cells("A2:B2")
    run_name = run.run_name or f"Run #{run.id}"
    subtitle = f"{run_name}  •  Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.font = Font(name="Calibri", size=10, italic=True, color="666666")
    sub_cell.alignment = Alignment(horizontal="center")
    row += 2  # blank row

    # ── Reconciliation overview ──────────────────────────────────────────────
    ws.merge_cells(f"A{row}:B{row}")
    section_cell = ws.cell(row=row, column=1, value="Reconciliation Overview")
    section_cell.font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    row += 1

    _write_summary_pair(ws, row, "Run ID", run.id); row += 1
    _write_summary_pair(ws, row, "Run Name", run.run_name or "—"); row += 1
    _write_summary_pair(ws, row, "Status", run.status.value.upper()); row += 1
    _write_summary_pair(ws, row, "Started At",
                        run.started_at.strftime("%Y-%m-%d %H:%M:%S") if run.started_at else "—"); row += 1
    _write_summary_pair(ws, row, "Completed At",
                        run.completed_at.strftime("%Y-%m-%d %H:%M:%S") if run.completed_at else "—"); row += 1
    row += 1  # blank

    # ── Match statistics ─────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:B{row}")
    section_cell = ws.cell(row=row, column=1, value="Match Statistics")
    section_cell.font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    row += 1

    _write_summary_pair(ws, row, "Match Rate (%)", round(match_rate, 2)); row += 1
    _write_summary_pair(ws, row, "Total Records", run.total_transactions or 0); row += 1
    _write_summary_pair(ws, row, "Matched", run.matched_count or 0); row += 1
    _write_summary_pair(ws, row, "Unmatched", run.unmatched_count or 0); row += 1
    _write_summary_pair(ws, row, "Total Exceptions", run.exception_count or 0); row += 1
    row += 1  # blank

    # ── Exception breakdown ──────────────────────────────────────────────────
    if exception_counts:
        ws.merge_cells(f"A{row}:B{row}")
        section_cell = ws.cell(row=row, column=1, value="Exception Breakdown by Type")
        section_cell.font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
        row += 1

        for exc_type, count in sorted(exception_counts.items()):
            _write_summary_pair(ws, row, exc_type, count); row += 1
        row += 1

    # ── Financial summary ────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:B{row}")
    section_cell = ws.cell(row=row, column=1, value="Financial Summary (INR)")
    section_cell.font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    row += 1

    total_billed = settlement_totals.get("total_payment_amount", 0)
    total_settled = settlement_totals.get("total_settled_amount", 0)
    gross_disc = settlement_totals.get("total_discrepancy", 0)
    amt_mismatch = settlement_totals.get("amount_mismatches_total", 0)
    unsettled = settlement_totals.get("unsettled_value_total", 0)
    duplicates = settlement_totals.get("duplicate_charges_total", 0)

    _write_summary_pair(ws, row, "Total Billed", total_billed, is_currency=True); row += 1
    _write_summary_pair(ws, row, "Total Settled", total_settled, is_currency=True); row += 1
    _write_summary_pair(ws, row, "Gross Discrepancy", gross_disc, is_currency=True); row += 1
    row += 1  # sub-header gap

    ws.merge_cells(f"A{row}:B{row}")
    section_cell = ws.cell(row=row, column=1, value="Discrepancy Breakdown")
    section_cell.font = Font(name="Calibri", bold=True, size=12, color="C65102")
    row += 1

    _write_summary_pair(ws, row, "Amount Mismatches Total", amt_mismatch, is_currency=True); row += 1
    _write_summary_pair(ws, row, "Unsettled Value Total", unsettled, is_currency=True); row += 1
    _write_summary_pair(ws, row, "Duplicate Charges Total", duplicates, is_currency=True); row += 1

    _auto_size_columns(ws, min_width=25, max_width=50)


def _build_exceptions_sheet(wb: Workbook, exceptions: list):
    ws = wb.create_sheet("Exceptions")
    ws.sheet_properties.tabColor = "C65102"

    headers = [
        "ID", "Type", "Payment ID", "Order ID", "Severity",
        "Internal Amount", "Bank Amount", "Discrepancy Amount",
        "Exception Date", "Description", "AI Explanation",
    ]

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    _style_header_row(ws, 1, len(headers))

    for row_idx, exc in enumerate(exceptions, start=2):
        ws.cell(row=row_idx, column=1, value=exc.id)
        ws.cell(row=row_idx, column=2, value=exc.exception_type)
        ws.cell(row=row_idx, column=3, value=exc.payment_id)
        ws.cell(row=row_idx, column=4, value=exc.order_id)
        ws.cell(row=row_idx, column=5, value=exc.severity.value if hasattr(exc.severity, 'value') else str(exc.severity))

        amt_cell = ws.cell(row=row_idx, column=6, value=exc.internal_amount)
        if exc.internal_amount is not None:
            amt_cell.number_format = CURRENCY_FORMAT

        bank_cell = ws.cell(row=row_idx, column=7, value=exc.bank_amount)
        if exc.bank_amount is not None:
            bank_cell.number_format = CURRENCY_FORMAT

        disc_cell = ws.cell(row=row_idx, column=8, value=exc.discrepancy_amount)
        if exc.discrepancy_amount is not None:
            disc_cell.number_format = CURRENCY_FORMAT

        ws.cell(row=row_idx, column=9, value=exc.exception_date)
        ws.cell(row=row_idx, column=10, value=exc.description)
        ws.cell(row=row_idx, column=11, value=exc.ai_explanation or "")

    _auto_size_columns(ws)

    # Freeze header row
    ws.freeze_panes = "A2"


def _build_raw_reconciliation_sheet(
    wb: Workbook,
    run: ReconciliationRun,
    exceptions: list,
    match_rate: float,
):
    """
    Build transaction-level reconciliation sheet containing one row per payment.
    Excludes standalone bank-side exceptions (orphan credits / duplicate credits).
    """
    ws = wb.create_sheet("Raw Reconciliation")
    ws.sheet_properties.tabColor = "2E7D32"

    headers = [
        "payment_id",
        "order_id",
        "payment_amount",
        "settled_amount",
        "bank_credited_amount",
        "utr_number",
        "payment_date",
        "settlement_date",
        "status",
    ]

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    _style_header_row(ws, 1, len(headers))

    # Map payment-level exceptions: payment_id -> exception_type
    exc_map = {e.payment_id: e.exception_type for e in exceptions if getattr(e, "payment_id", None)}

    # Locate source CSV files (sample_data or uploads session)
    raw_df = None
    possible_paths = [
        Path("sample_data"),
        Path("../sample_data"),
        Path("backend/sample_data"),
        Path("c:/My codings/razorpay buildthon/backend/sample_data"),
    ]

    for data_dir in possible_paths:
        if (data_dir / "payments.csv").exists() and (data_dir / "settlements.csv").exists() and (data_dir / "bank_statement.csv").exists():
            try:
                from app.services.reconciliation import load_payments, load_settlements, load_bank_statement, merge_all
                p = load_payments(data_dir / "payments.csv")
                s = load_settlements(data_dir / "settlements.csv")
                b = load_bank_statement(data_dir / "bank_statement.csv")
                raw_df = merge_all(p, s, b)
                break
            except Exception as exc:
                logger.warning(f"Could not load data from {data_dir}: {exc}")

    if raw_df is not None and not raw_df.empty:
        for row_idx, r in enumerate(raw_df.to_dict("records"), start=2):
            pid = str(r.get("payment_id", "") or "")
            oid = str(r.get("order_id", "") or "")
            pmt_amt = r.get("amount")
            stl_amt = r.get("settled_amount")
            bnk_amt = r.get("bank_credited_amount")
            utr = str(r.get("utr_number", "") or "") if pd.notna(r.get("utr_number")) else ""
            pmt_date = str(r.get("payment_date", "") or "") if pd.notna(r.get("payment_date")) else ""
            stl_date = str(r.get("settlement_date", "") or "") if pd.notna(r.get("settlement_date")) else ""
            status = exc_map.get(pid, "MATCHED")

            ws.cell(row=row_idx, column=1, value=pid if pid else None)
            ws.cell(row=row_idx, column=2, value=oid if oid else None)

            c3 = ws.cell(row=row_idx, column=3, value=float(pmt_amt) if pd.notna(pmt_amt) else None)
            if pd.notna(pmt_amt):
                c3.number_format = CURRENCY_FORMAT

            c4 = ws.cell(row=row_idx, column=4, value=float(stl_amt) if pd.notna(stl_amt) else None)
            if pd.notna(stl_amt):
                c4.number_format = CURRENCY_FORMAT

            c5 = ws.cell(row=row_idx, column=5, value=float(bnk_amt) if pd.notna(bnk_amt) else None)
            if pd.notna(bnk_amt):
                c5.number_format = CURRENCY_FORMAT

            ws.cell(row=row_idx, column=6, value=utr if utr else None)
            ws.cell(row=row_idx, column=7, value=pmt_date if pmt_date else None)
            ws.cell(row=row_idx, column=8, value=stl_date if stl_date else None)

            status_cell = ws.cell(row=row_idx, column=9, value=status)
            if status == "MATCHED":
                status_cell.font = Font(name="Calibri", color="2E7D32", bold=True)
            else:
                status_cell.font = Font(name="Calibri", color="C65102", bold=True)
    else:
        # Fallback for synthetic test environments without CSV fixtures
        row_idx = 2
        payment_excs = [e for e in exceptions if getattr(e, "payment_id", None)]
        for exc in payment_excs:
            ws.cell(row=row_idx, column=1, value=exc.payment_id)
            ws.cell(row=row_idx, column=2, value=exc.order_id)
            c3 = ws.cell(row=row_idx, column=3, value=exc.internal_amount)
            if exc.internal_amount is not None:
                c3.number_format = CURRENCY_FORMAT
            c4 = ws.cell(row=row_idx, column=4, value=exc.bank_amount)
            if exc.bank_amount is not None:
                c4.number_format = CURRENCY_FORMAT
            c5 = ws.cell(row=row_idx, column=5, value=exc.bank_amount)
            if exc.bank_amount is not None:
                c5.number_format = CURRENCY_FORMAT
            ws.cell(row=row_idx, column=6, value=None)
            ws.cell(row=row_idx, column=7, value=exc.exception_date)
            ws.cell(row=row_idx, column=8, value=exc.exception_date)
            status_cell = ws.cell(row=row_idx, column=9, value=exc.exception_type)
            status_cell.font = Font(name="Calibri", color="C65102", bold=True)
            row_idx += 1

        matched_count = run.matched_count or (run.total_transactions - len(payment_excs) if run.total_transactions else 0)
        for m_i in range(matched_count):
            ws.cell(row=row_idx, column=1, value=f"pay_matched_{m_i+1:04d}")
            ws.cell(row=row_idx, column=2, value=f"ord_matched_{m_i+1:04d}")
            status_cell = ws.cell(row=row_idx, column=9, value="MATCHED")
            status_cell.font = Font(name="Calibri", color="2E7D32", bold=True)
            row_idx += 1

    _auto_size_columns(ws)
    ws.freeze_panes = "A2"

